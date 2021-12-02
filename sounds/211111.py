# '''读取数据'''
import numpy
import numpy as np
import matplotlib.pyplot as plt  # 导入绘图工作的函数集合
from openpyxl import load_workbook
import pandas as pd
import operator


# 获取Excel数据
# wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1102数据/噪音/噪音1.xlsx')
wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1102数据/组件+多余物/组多2.xlsx')
# wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1021数据/组件1.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

# 获取第一列所有数据
col1 = []
for col in sheet1['A'][1:200000]:
    col1.append(col.value)
col2 = []
for col in sheet1['B'][1:200000]:
    col2.append(col.value)

col3 = []
for col in sheet1['C'][1:200000]:
    col3.append(col.value)

col4 = []
for col in sheet1['D'][1:200000]:
    col4.append(col.value)

col5 = []
for col in sheet1['E'][1:200000]:
    col5.append(col.value)

col6 = []
for col in sheet1['F'][1:200000]:
    col6.append(col.value)

col7 = []
for col in sheet1['G'][1:200000]:
    col7.append(col.value)

col8 = []
for col in sheet1['H'][1:200000]:
    col8.append(col.value)

col9 = []
for col in sheet1['I'][1:200000]:
    col9.append(col.value)

col10 = []
for col in sheet1['J'][1:200000]:
    col10.append(col.value)

col11 = []
for col in sheet1['K'][1:200000]:
    col11.append(col.value)

col12 = []
for col in sheet1['L'][1:200000]:
    col12.append(col.value)

col13 = []
for col in sheet1['M'][1:200000]:
    col13.append(col.value)

col14 = []
for col in sheet1['N'][1:200000]:
    col4.append(col.value)

col15 = []
for col in sheet1['O'][1:200000]:
    col15.append(col.value)

col16 = []
sum = 0
Mean = 0.0

# -----------------计算每一列平均值-------------------
for i in range(2, 16, 1):
   for j in range(0, 199999, 1):  # 循环逐行打印
      m = coli[j]  # 获取第2列的数字
      n = abs(m)   # 绝对值
      sum += n
      Mean = sum / len(coli)  # 计算平均能量
    print(Mean)