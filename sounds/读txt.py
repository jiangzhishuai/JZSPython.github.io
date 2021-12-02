import pandas as pd
import numpy
import numpy as np
import matplotlib.pyplot as plt  # 导入绘图工作的函数集合
from openpyxl import load_workbook
import operator

# -------------------  读取txt数据 存储到数组data里-------------------------

with open("C:/Users/zhangming/Desktop/1119下午/无多余物/1.txt", "r") as f:  # 打开文件
   # data = f.read().splitlines()   # 去除换行符号
   data = f.readlines()             # 直接出去  不去除\n
   f.close()
print(type(data))
print(len(data))
# print(data)

B = np.array(data)
print(B)

C = B[0:100]

data = pd.DataFrame(C)
writer = pd.ExcelWriter('C:/Users/zhangming/Desktop/1119下午/无多余物/test31.xlsx')
# header参数表示列的名称，index表示行的标签
data.to_excel(writer, 'sheet_1', float_format='%.6f', header=False, index=False)
writer.save()
writer.close()

wb = load_workbook('C:/Users/zhangming/Desktop/1119下午/无多余物/test31.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

# 获取第一列所有数据
col2 = []
for col in sheet1['A'][0:100]:
    col2.append(float(col.value))
D = col2


# -------------------  读取#EXCE 数据 存储到数组time横坐标里-------------------------
wb = load_workbook('C:/Users/zhangming/Desktop/1119下午/time.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

# 获取第一列所有数据
col1 = []
for col in sheet1['A'][0:100]:
    col1.append(col.value)
A = col1


# 读取音频文件
y = D
x = A  # 时间刻度
# 绘制图形
plt.plot(x, y)
plt.xlabel('t/s')  # x轴时间
plt.ylabel('U/V')  # y轴振幅
plt.title('shiyutu', fontsize=12, color='black')  # 标题名称、字体大小、颜色
plt.ylim(-10, 10)
plt.show()

