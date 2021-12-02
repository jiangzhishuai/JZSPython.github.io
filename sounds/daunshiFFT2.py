import numpy as np
from openpyxl import load_workbook
from scipy import signal
import matplotlib.pyplot as plt

# --------------获取Excel数据-----------------
# wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1021数据/多余物2.xlsx')
wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1102数据/组件+多余物/组多2.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
print(sheets)

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

# 获取第一列所有数据  存到数组col1中  时间
col1 = []
for col in sheet1['A'][0:200000]:
    col1.append(col.value)

# 获取第二列所有数据  存到数组col2中  电压
col2 = []
for col in sheet1['B'][0:200000]:
    col2.append(col.value)

# --------------短时傅里叶变换----------------
fs = 200000
N = 200000
time = np.arange(N) / float(fs)
x = col2
f, t, Zxx = signal.stft(x, fs, nperseg=1000)
plt.pcolormesh(t, f, np.abs(Zxx), shading='gouraud')
plt.title('STFT Magnitude')  # 图片标题
plt.ylabel('Frequency [Hz]')  # 纵坐标
plt.xlabel('Time [sec]')  # 横坐标
# plt.show()
# print(signal.stft(x, fs, nperseg=1000))
print(Zxx)