# '''读取数据'''
import numpy
import numpy as np
import matplotlib.pyplot as plt  # 导入绘图工作的函数集合
from openpyxl import load_workbook
import pandas as pd
import operator


# 获取Excel数据
# wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1102数据/噪音/噪音1.xlsx')
wb = load_workbook('C:/Users/zhangming/Desktop/1110数据2/1111简化.xlsx')
# wb = load_workbook('C:/Users/zhangming/Desktop/采集卡1210/1021数据/组件1.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

# 获取第一列所有数据
col1 = []
for col in sheet1['A'][0:200000]:
    col1.append(col.value)
col2 = []
for col in sheet1['O'][0:200000]:
    col2.append(col.value)
col3 = []
col4 = []
col5 = []
col6 = []
col7 = []
col9 = []

sum = 0
for j in range(len(col2)):  # 循环逐行打印
    m = col2[j]  # 获取第2列的数字
    n = abs(m)   # 绝对值
    sum = sum + n

Umean = sum / len(col2)  # 计算平均能量
print(Umean)

Upeak = Umean * 8  # 计算脉冲峰值阈值
Ustart = Umean * 1.2  # 计算脉冲开始阈值
Uend = Umean * 1.1  # 计算脉冲结束阈值

print("平均能量：" + '%.6f' % Umean)
print("脉冲峰值阈值："+ '%.6f' % Upeak)      # 脉冲峰值阈值
print("脉冲开始阈值："+ '%.6f' % Ustart)  # 脉冲开始阈值
print()

tss = []  # 脉冲开始时间数组 col2.append(col.value)
ss = len(col2)
sa = 0
ge = 0
jg1 = 20
jg2 = 10
tend = 0
# 打印脉冲时间
for a in range(0, ss, jg1):  # 循环逐行
    if a > sa:  # 保证下一个脉冲从是一个脉冲结束时间之后找
        if col2[a] >= Upeak:
            Tmax1 = col1[a]    # 峰值时间
            Tu = col2[a]       # 峰值
            print("峰值时间：" + '%.6f' % Tmax1+"峰值："+ '%.6f' % Tu)
            for q in range(a, sa, -jg2):  # 往前找开始脉冲时间
                if 0 < col2[q] <= Ustart:
                    Tstart = col1[q]
                    Z1 = q
                    print("开始时间：" + '%.6f' % Tstart)
                    tss.append(Tstart)
                    break
            for p in range(a, ss, jg2):  # 往后找结束脉冲时间
                if 0 < col2[p] <= Uend:
                    Tend = col1[p]
                    Z2 = p
                    print("结束时间：" + '%.6f' % Tend)
                    sa = p
                    break
            ge += 1

print("提取脉冲得个数为："+ '%.2f' % ge)