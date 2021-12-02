import math
import time
import pandas as pd
import numpy
import numpy as np
import matplotlib.pyplot as plt  # 导入绘图工作的函数集合
from openpyxl import load_workbook
from scipy.fftpack import fft
import matplotlib.pyplot as plt
from matplotlib.pylab import mpl

# -------------------  读取txt数据 存储到数组data里-------------------------
# with open("C:/Users/zhangming/Desktop/1119下午/无多余物/1.txt", "r") as f:  # 打开文件
# with open("C:/Users/zhangming/Desktop/1122上午/11.22/线/19.txt", "r") as f:  # 打开文件
# with open("C:/Users/zhangming/Desktop/采集数据/1124晚上一/线/  (15).txt", "r") as f:  # 打开文件
with open("D:/Jiangshuai/研二/采集数据/1201/无/(1).txt", "r") as f:  # 打开文件
   # data = f.read().splitlines()   # 去除换行符号
   data = f.readlines()             # 直接出去  不去除\n
   f.close()
B = np.array(data)  #  list 转 narr

col2 = []
# 遍历数组 转化为float
for i in B[2:len(B)]:
    col2.append(float(i))
print(len(col2))

# print(D)


# -------------------  读取#EXCE 数据 存储到数组time横坐标里-------------------------
# wb = load_workbook('C:/Users/zhangming/Desktop/1119下午/time2.xlsx')
wb = load_workbook('C:/Users/zhangming/Desktop/time.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet

# 获取第一张sheet
sheet1 = sheets[0]
print(sheet1)

# 获取第一列所有数据
col1 = []
for col in sheet1['A'][0:100000]:
    col1.append(col.value)
A = np.array(col1)
# C = col2[0:100000]
C = col2

# 频率特征值
def get_fre_domain_features(f, y):


   fre_line_num = len(y)

   p1 = y.mean()
   p2 = math.sqrt(sum((y - p1) ** 2) / fre_line_num)
   p3 = sum((y - p1) ** 3) / (fre_line_num * p2 ** 3)
   p4 = sum((y - p1) ** 4) / (fre_line_num * p2 ** 4)
   p5 = sum(f * y) / sum(y)
   p6 = math.sqrt(sum((f - p5) ** 2 * y) / fre_line_num)
   p7 = math.sqrt(sum(f ** 2 * y) / sum(y))
   # p8 = math.sqrt(sum(f ** 4 * y) / sum(f ** 2 * y))
   # p9 = sum(f ** 2 * y) / math.sqrt(sum(y) * sum(f ** 4 * y))
   p10 = p6 / p5
   p11 = sum((f - p5) ** 3 * y) / (p6 ** 3 * fre_line_num)
   p12 = sum((f - p5) ** 4 * y) / (p6 ** 4 * fre_line_num)
   p13 = sum(abs(f - p5) * y) / (math.sqrt(p6) * fre_line_num)
   p = [p1, p2, p3, p4, p5, p6, p7,  p10, p11, p12, p13]
   return p

# 插件txt文件并写入数据
def text_create(name, msg):

   desktop_path = "C:/Users/zhangming/Desktop/" # 新创建的txt文件的存放路径
   full_path = desktop_path + name + '.txt'  # 也可以创建一个.doc的word文档
   file = open(full_path, 'w')
   file.write(msg)  # msg也就是下面的Hello world!


# file.close()


# 绘图函数
def huitu( A, B, c ):  # (时间，电压)

   plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=1)  #  调节各个图片之间的距离，使显示更加合理


   mpl.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文
   mpl.rcParams['axes.unicode_minus'] = False  # 显示负号
   # 读取音频文件
   y = B
   x = A  # 时间刻度
   # 绘制图形
   plt.subplot(311)
   plt.plot(x, y)
   plt.xlabel('t/s')  # x轴时间
   plt.ylabel('U/V')  # y轴振幅
   plt.title('时域图', fontsize=12, color='black')  # 标题名称、字体大小、颜色
   plt.ylim(-10, 10)
   # plt.show()

   # -------------------  傅里叶变换 -------------------------

   N = 100000
   A = np.array(col1)  # LIST转ARRRY
   x = A
   y = B

   fft_y = fft(y)  # 快速傅里叶变换

   x = np.arange(N)  # 频率个数
   half_x = x[range(int(N / 2))]  # 取一半区间

   angle_y = np.angle(fft_y)  # 取复数的角度

   abs_y = np.abs(fft_y)  # 取复数的绝对值，即复数的模(双边频谱)
   normalization_y = abs_y / (N / 2)  # 归一化处理（双边频谱）
   normalization_y[0] /= 2  # 归一化处理（双边频谱）
   normalization_half_y = normalization_y[range(int(N / 2))]  # 由于对称性，只取一半区间（单边频谱）


   """
   plt.subplot(242)
   plt.xlim(0, 50000)
   plt.xlabel('Frequency/Hz')  # x轴频率
   plt.ylabel('幅度/dB')  # y轴幅度
   plt.plot(x, y)
   plt.title('原始波形')
    """

   """
   plt.subplot(243)
   plt.xlim(0, 50000)
   plt.xlabel('Frequency/Hz')  # x轴频率
   plt.ylabel('幅度/dB')  # y轴幅度
   plt.plot(x, fft_y, 'black')
   plt.title('双边振幅谱(未求振幅绝对值)', fontsize=9, color='black')
    """

   plt.subplot(312)
   plt.xlim(0, 50000)
   plt.xlabel('Frequency/Hz')  # x轴频率
   plt.ylabel('幅度/dB')  # y轴幅度
   plt.plot(x, abs_y, 'r')
   plt.ylim(0, 500)
   plt.title('单边振幅谱(未归一化)', fontsize=9, color='red')

   """
   plt.subplot(245)
   plt.xlim(0, 50000)
   plt.xlabel('Frequency/Hz')  # x轴频率
   plt.ylabel('幅度/dB')  # y轴幅度
   plt.plot(x, angle_y, 'violet')
   plt.title('双边相位谱(未归一化)', fontsize=9, color='violet')
   

   plt.subplot(246)
   plt.xlim(0, 50000)
   plt.xlabel('Frequency/Hz')  # x轴频率
   plt.ylabel('幅度/dB')  # y轴幅度
   plt.plot(x, normalization_y, 'g')
   plt.title('双边振幅谱(归一化)', fontsize=9, color='green')
   """

   plt.subplot(313)
   plt.xlim(0, 50000)
   plt.xlabel('Frequency/Hz')  # x轴频率
   plt.ylabel('幅度/dB')  # y轴幅度
   plt.plot(half_x, normalization_half_y, 'blue')
   plt.title('单边振幅谱(归一化)', fontsize=9, color='blue')
   # plt.show()


   x2 = x[0:50000]
   y2 = abs_y[0:50000]

   """
   print(x)
   print(len(x))
   print(type(x))
   print(abs_y)
   print(len(abs_y))
   print(type(abs_y))
   """

   print(get_fre_domain_features(x2,y2))


   now = time.strftime("%Y-%m-%d-%H_%M_%S", time.localtime(time.time()))
   c = str(c)


   data = pd.DataFrame(get_fre_domain_features(x2,y2))
   writer = pd.ExcelWriter('C:/Users/zhangming/Desktop/test' + c + '.xlsx')
   # header参数表示列的名称，index表示行的标签
   data.to_excel(writer, 'sheet_1', float_format='%.6f', header=False, index=False)
   writer.save()
   writer.close()





   # return get_fre_domain_features(x2, y2)


# 使得几秒的数据分段显示
a = len(col2) // 100000
print(a)

"""
F = []
for b in range(0, a, 1):  # 循环逐行
    D = col2[b * 100000: b * 100000 + 100000]
    huitu(A, D, b)
"""


#  第4秒数据最典型，所以进行观看，分析
D = col2[(a-2) * 100000: (a-2) * 100000 + 100000]
huitu(A, D, 2)
